from flask import Blueprint, render_template, redirect, url_for, send_file, make_response
from flask_login import login_required, current_user
from models.transaction import Transaction
from models.savings_goal import SavingsGoal
from models.category import Category
from models.monthly_budget import MonthlyBudget
from app import db
from sqlalchemy import func, extract
from datetime import datetime, timedelta
import calendar
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

main_bp = Blueprint('main', __name__)

@main_bp.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    return redirect(url_for('auth.login'))

@main_bp.route('/dashboard')
@login_required
def dashboard():
    # Get current date info
    today = datetime.now().date()
    current_month = today.month
    current_year = today.year
    
    # Calculate date ranges
    month_start = today.replace(day=1)
    if current_month == 12:
        month_end = today.replace(year=current_year + 1, month=1, day=1) - timedelta(days=1)
    else:
        month_end = today.replace(month=current_month + 1, day=1) - timedelta(days=1)
    
    # Get user's transactions
    user_transactions = Transaction.query.filter_by(user_id=current_user.id)
    
    # Calculate totals
    total_income = user_transactions.filter_by(type='income').with_entities(func.sum(Transaction.amount)).scalar() or 0
    total_expense = user_transactions.filter_by(type='expense').with_entities(func.sum(Transaction.amount)).scalar() or 0
    balance = total_income - total_expense
    
    # Monthly totals
    monthly_income = user_transactions.filter(
        Transaction.type == 'income',
        Transaction.date >= month_start,
        Transaction.date <= month_end
    ).with_entities(func.sum(Transaction.amount)).scalar() or 0
    
    monthly_expense = user_transactions.filter(
        Transaction.type == 'expense',
        Transaction.date >= month_start,
        Transaction.date <= month_end
    ).with_entities(func.sum(Transaction.amount)).scalar() or 0
    
    # Recent transactions
    recent_transactions = user_transactions.order_by(Transaction.created_at.desc()).limit(5).all()
    
    # Category spending this month
    category_spending = db.session.query(
        Category.name,
        func.sum(Transaction.amount).label('total')
    ).join(Transaction).filter(
        Transaction.user_id == current_user.id,
        Transaction.type == 'expense',
        Transaction.date >= month_start,
        Transaction.date <= month_end
    ).group_by(Category.name).order_by(func.sum(Transaction.amount).desc()).limit(5).all()
    
    # Savings goals
    savings_goals = SavingsGoal.query.filter_by(user_id=current_user.id, is_active=True).all()
    
    # Monthly comparison data for chart
    monthly_data = []
    for i in range(6):  # Last 6 months
        target_date = today.replace(day=1) - timedelta(days=i*30)
        month_income = user_transactions.filter(
            Transaction.type == 'income',
            extract('month', Transaction.date) == target_date.month,
            extract('year', Transaction.date) == target_date.year
        ).with_entities(func.sum(Transaction.amount)).scalar() or 0
        
        month_expense = user_transactions.filter(
            Transaction.type == 'expense',
            extract('month', Transaction.date) == target_date.month,
            extract('year', Transaction.date) == target_date.year
        ).with_entities(func.sum(Transaction.amount)).scalar() or 0
        
        monthly_data.append({
            'month': calendar.month_name[target_date.month],
            'income': float(month_income),
            'expense': float(month_expense)
        })
    
    monthly_data.reverse()
    
    # Budget Alert Information
    budget_alert = None
    current_budget = MonthlyBudget.get_current_month_budget(current_user.id)
    if current_budget and current_budget.budget_limit > 0:
        spending_percentage = (float(monthly_expense) / float(current_budget.budget_limit)) * 100
        
        # Chỉ hiển thị alert khi >= 70% (thay vì 80% để hiển thị sớm hơn)
        if spending_percentage >= 70:
            # Xác định mức độ cảnh báo
            if spending_percentage >= 100:
                alert_level = 'danger'
                alert_color = 'danger'
                alert_message = 'Đã vượt quá giới hạn chi tiêu!'
                alert_title = '🚨 Vượt Giới Hạn!'
            elif spending_percentage >= 95:
                alert_level = 'critical'
                alert_color = 'danger'
                alert_message = 'Sắp vượt quá giới hạn chi tiêu!'
                alert_title = '⚠️ Nguy Hiểm!'
            elif spending_percentage >= 80:
                alert_level = 'warning'
                alert_color = 'warning'
                alert_message = 'Đã chi tiêu gần đạt giới hạn tháng'
                alert_title = '⚡ Cảnh Báo!'
            else:  # 70-80%
                alert_level = 'info'
                alert_color = 'info'
                alert_message = 'Chi tiêu đang tăng, cần chú ý'
                alert_title = '📊 Theo Dõi'
            
            budget_alert = {
                'budget_limit': float(current_budget.budget_limit),
                'current_spending': float(monthly_expense),
                'remaining_budget': float(current_budget.budget_limit) - float(monthly_expense),
                'spending_percentage': round(spending_percentage, 1),
                'alert_level': alert_level,
                'alert_color': alert_color,
                'alert_message': alert_message,
                'alert_title': alert_title,
                'show_alert': True
            }
    
    return render_template('main/dashboard.html',
                         total_income=total_income,
                         total_expense=total_expense,
                         balance=balance,
                         monthly_income=monthly_income,
                         monthly_expense=monthly_expense,
                         recent_transactions=recent_transactions,
                         category_spending=category_spending,
                         savings_goals=savings_goals,
                         monthly_data=monthly_data,
                         budget_alert=budget_alert)

@main_bp.route('/profile')
@login_required
def profile():
    return render_template('main/profile.html')

@main_bp.route('/export/transactions')
@login_required
def export_transactions():
    """Export all user transactions to Excel file"""
    try:
        # Lấy tất cả giao dịch của user
        transactions = Transaction.query.filter_by(user_id=current_user.id)\
                                      .order_by(Transaction.date.desc())\
                                      .all()
        
        # Tạo workbook và worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Giao Dịch"
        
        # Định nghĩa headers
        headers = [
            'STT', 'Ngày', 'Loại', 'Danh mục', 'Mô tả', 
            'Số tiền (VNĐ)', 'Hóa đơn', 'Ngày tạo'
        ]
        
        # Style cho header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Thêm headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Thêm dữ liệu
        for row, transaction in enumerate(transactions, 2):
            # STT
            ws.cell(row=row, column=1, value=row-1).border = border
            
            # Ngày
            date_cell = ws.cell(row=row, column=2, value=transaction.date.strftime('%d/%m/%Y'))
            date_cell.border = border
            
            # Loại
            type_text = "Thu nhập" if transaction.type == 'income' else "Chi tiêu"
            type_cell = ws.cell(row=row, column=3, value=type_text)
            type_cell.border = border
            # Màu sắc theo loại
            if transaction.type == 'income':
                type_cell.font = Font(color="008000")  # Xanh lá
            else:
                type_cell.font = Font(color="FF0000")  # Đỏ
            
            # Danh mục
            category_cell = ws.cell(row=row, column=4, value=transaction.category.name if transaction.category else "N/A")
            category_cell.border = border
            
            # Mô tả
            desc_cell = ws.cell(row=row, column=5, value=transaction.description or "")
            desc_cell.border = border
            
            # Số tiền
            amount_cell = ws.cell(row=row, column=6, value=float(transaction.amount))
            amount_cell.number_format = '#,##0'
            amount_cell.border = border
            amount_cell.alignment = Alignment(horizontal="right")
            
            # Hóa đơn
            receipt_cell = ws.cell(row=row, column=7, value="Có" if transaction.receipt_image else "Không")
            receipt_cell.border = border
            
            # Ngày tạo
            created_cell = ws.cell(row=row, column=8, value=transaction.created_at.strftime('%d/%m/%Y %H:%M'))
            created_cell.border = border
        
        # Tự động điều chỉnh độ rộng cột
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Thêm thống kê ở cuối
        if transactions:
            stats_row = len(transactions) + 3
            
            # Tổng thu nhập
            total_income = sum(t.amount for t in transactions if t.type == 'income')
            ws.cell(row=stats_row, column=5, value="Tổng thu nhập:").font = Font(bold=True)
            income_cell = ws.cell(row=stats_row, column=6, value=float(total_income))
            income_cell.number_format = '#,##0'
            income_cell.font = Font(bold=True, color="008000")
            
            # Tổng chi tiêu
            total_expense = sum(t.amount for t in transactions if t.type == 'expense')
            ws.cell(row=stats_row + 1, column=5, value="Tổng chi tiêu:").font = Font(bold=True)
            expense_cell = ws.cell(row=stats_row + 1, column=6, value=float(total_expense))
            expense_cell.number_format = '#,##0'
            expense_cell.font = Font(bold=True, color="FF0000")
            
            # Số dư
            balance = total_income - total_expense
            ws.cell(row=stats_row + 2, column=5, value="Số dư:").font = Font(bold=True)
            balance_cell = ws.cell(row=stats_row + 2, column=6, value=float(balance))
            balance_cell.number_format = '#,##0'
            balance_cell.font = Font(bold=True, color="0000FF")
        
        # Lưu file vào memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Tạo tên file với timestamp
        filename = f"giao_dich_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Tạo response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        # Log error và redirect về dashboard với thông báo lỗi
        print(f"Export error: {str(e)}")
        from flask import flash
        flash(f'Lỗi khi export Excel: {str(e)}', 'error')
        return redirect(url_for('main.dashboard'))